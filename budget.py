import openpyxl
import curses
import os

# Initialize Budget File
def initialize_budget_file(file_name="Household_Budget.xlsx"):
    if not os.path.exists(file_name):
        wb = openpyxl.Workbook()

        # Overview Sheet
        wb.active.title = "Settings"
        settings_sheet = wb.active
        settings_sheet.append(["Yearly Salary", 0])
        settings_sheet.append(["Savings Ratio", 0.5])  # Default: 50%
        settings_sheet.append(["Checking Ratio", 0.5])  # Default: 50%

        # Expenses Sheet
        expenses_sheet = wb.create_sheet("Expenses")
        expenses_sheet.append(["Type", "Category", "Description", "Amount", "Frequency (weekly/monthly/yearly)"])

        wb.save(file_name)
        print(f"New budget file '{file_name}' created.")
    else:
        print(f"Loading existing budget file '{file_name}'.")


# Prompt for user input outside curses
def get_input(stdscr, prompt_text):
    # Suspend curses screen
    curses.nocbreak()
    stdscr.keypad(False)
    curses.echo()
    stdscr.refresh()
    
    # Prompt for input
    stdscr.clear()
    stdscr.addstr(prompt_text)
    stdscr.refresh()
    user_input = stdscr.getstr().decode("utf-8")
    
    # Resume curses screen
    curses.noecho()
    curses.cbreak()
    stdscr.keypad(True)
    return user_input


# Add or Adjust Yearly Salary
def set_yearly_salary(file_name, salary):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb["Settings"]
    sheet["B1"] = salary
    wb.save(file_name)
    print(f"Yearly salary set to ${salary}.")

# Set Savings/Checking Ratios
def set_ratios(file_name, savings_ratio):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb["Settings"]
    checking_ratio = 1 - savings_ratio
    sheet["B2"] = savings_ratio
    sheet["B3"] = checking_ratio
    wb.save(file_name)
    print(f"Savings ratio set to {savings_ratio * 100}%. Checking ratio set to {checking_ratio * 100}%.")

# Add Expense
def add_expense(file_name, exp_type, category, description, amount, frequency):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb["Expenses"]
    sheet.append([exp_type, category, description, amount, frequency])
    wb.save(file_name)
    print(f"Expense '{description}' added.")

# Function to load expenses from the Excel file
def load_expenses(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb["Expenses"]
    expenses = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        expenses.append(row)
    return expenses

# Function to delete an expense (updated to be used in curses menu)
def remove_expense(file_name, stdscr):
    expenses = load_expenses(file_name)

    # Function to draw the expense menu with pagination
    def draw_expense_menu(stdscr, current_idx, page_idx, expenses):
        stdscr.clear()
        stdscr.addstr("Select Expense to Remove (or Cancel):\n", curses.A_BOLD | curses.A_UNDERLINE)

        # Paginate the expenses
        start_idx = page_idx * 10
        end_idx = min(start_idx + 10, len(expenses))
        expenses_to_display = expenses[start_idx:end_idx]

        # Display expenses
        for idx, expense in enumerate(expenses_to_display):
            expense_str = f"{expense[0]} - {expense[1]} - {expense[2]} - ${expense[3]:.2f}"
            if idx == current_idx:
                stdscr.addstr(f"> {expense_str}\n", curses.A_REVERSE)
            else:
                stdscr.addstr(f"  {expense_str}\n")

        # Add the "Cancel" option
        if current_idx == len(expenses_to_display):
            stdscr.addstr("> Cancel\n", curses.A_REVERSE)
        else:
            stdscr.addstr("  Cancel\n")

        # Page navigation instructions
        if page_idx > 0:
            stdscr.addstr(f"Page {page_idx + 1}/{(len(expenses) // 10) + 1} - Left/Right to navigate\n")
        else:
            stdscr.addstr(f"Page 1/{(len(expenses) // 10) + 1} - Right to navigate\n")

        stdscr.refresh()

    # Initial setup
    current_idx = 0
    page_idx = 0

    while True:
        draw_expense_menu(stdscr, current_idx, page_idx, expenses)
        key = stdscr.getch()

        if key == curses.KEY_UP and current_idx > 0:
            current_idx -= 1
        elif key == curses.KEY_DOWN and current_idx < len(expenses[:(page_idx + 1) * 10]):
            current_idx += 1
        elif key == curses.KEY_LEFT and page_idx > 0:
            page_idx -= 1
            current_idx = 0
        elif key == curses.KEY_RIGHT and (page_idx + 1) * 10 < len(expenses):
            page_idx += 1
            current_idx = 0
        elif key in (curses.KEY_ENTER, 10, 13, 32):  # Enter or Space
            # If user selects "Cancel"
            if current_idx == len(expenses[:(page_idx + 1) * 10]):
                stdscr.clear()
                stdscr.addstr("No expense was removed.\n")
                stdscr.refresh()
                stdscr.getch()
                break

            # Remove the selected expense
            row_number = page_idx * 10 + current_idx + 2  # +2 to account for header and zero-based index
            wb = openpyxl.load_workbook(file_name)
            sheet = wb["Expenses"]
            if row_number < 2 or row_number > sheet.max_row:
                stdscr.addstr("Invalid row number.\n")
            else:
                sheet.delete_rows(row_number)
                wb.save(file_name)
                stdscr.clear()
                stdscr.addstr(f"Expense {expenses[current_idx]} removed.\n")
                stdscr.refresh()
                stdscr.getch()
                break
# Load unique categories from the Expenses sheet
def load_unique_categories(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb["Expenses"]
    categories = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        categories.add(row[1])  # Category is in column 2
    return list(categories)

# Add expenses to the file
def add_expense_menu(file_name, stdscr):
    unique_categories = load_unique_categories(file_name)

    # Expense Type Menu (Recurring/One-time)
    def draw_exp_type_menu(stdscr, current_idx):
        stdscr.clear()
        stdscr.addstr("Select Expense Type:\n", curses.A_BOLD | curses.A_UNDERLINE)
        options = ["Recurring", "One-time"]
        for idx, option in enumerate(options):
            if idx == current_idx:
                stdscr.addstr(f"> {option}\n", curses.A_REVERSE)
            else:
                stdscr.addstr(f"  {option}\n")
        stdscr.refresh()

# Category Selection Menu with Pagination
    def draw_category_menu(stdscr, current_idx, categories, page_idx):
        stdscr.clear()
        stdscr.addstr("Select Category (or Create New):\n", curses.A_BOLD | curses.A_UNDERLINE)

        # Display categories for the current page
        start_idx = page_idx * 10
        end_idx = min(start_idx + 10, len(categories))
        categories_to_display = categories[start_idx:end_idx] + ["New Category"]

        for idx, category in enumerate(categories_to_display):
            if idx == current_idx:
                stdscr.addstr(f"> {category}\n", curses.A_REVERSE)
            else:
                stdscr.addstr(f"  {category}\n")
        
        # Display page navigation instructions
        if page_idx > 0:
            stdscr.addstr(f"Page {page_idx + 1}/{(len(categories) // 10) + 1} - Left/Right to navigate\n")
        else:
            stdscr.addstr(f"Page 1/{(len(categories) // 10) + 1} - Right to navigate\n")

        stdscr.refresh()



    # Frequency Menu for Recurring Expenses
    def draw_frequency_menu(stdscr, current_idx):
        stdscr.clear()
        stdscr.addstr("Select Frequency for Recurring Expense:\n", curses.A_BOLD | curses.A_UNDERLINE)
        options = ["Weekly", "Monthly", "Yearly"]
        for idx, option in enumerate(options):
            if idx == current_idx:
                stdscr.addstr(f"> {option}\n", curses.A_REVERSE)
            else:
                stdscr.addstr(f"  {option}\n")
        stdscr.refresh()

    # Selecting Expense Type
    exp_type = None
    current_idx = 0
    while exp_type is None:
        draw_exp_type_menu(stdscr, current_idx)
        key = stdscr.getch()
        if key == curses.KEY_UP and current_idx > 0:
            current_idx -= 1
        elif key == curses.KEY_DOWN and current_idx < 1:
            current_idx += 1
        elif key in (curses.KEY_ENTER, 10, 13, 32):  # Enter or Space
            exp_type = ["Recurring", "One-time"][current_idx]

    # Selecting Category with Pagination
    category = None
    current_idx = 0
    page_idx = 0
    while category is None:
        draw_category_menu(stdscr, current_idx, unique_categories, page_idx)
        key = stdscr.getch()

        if key == curses.KEY_UP and current_idx > 0:
            current_idx -= 1
        elif key == curses.KEY_DOWN and current_idx < len(unique_categories):
            current_idx += 1
        elif key == curses.KEY_LEFT and page_idx > 0:
            page_idx -= 1
            current_idx = 0  # Reset to the first item of the new page
        elif key == curses.KEY_RIGHT and (page_idx + 1) * 10 < len(unique_categories):
            page_idx += 1
            current_idx = 0  # Reset to the first item of the new page
        elif key in (curses.KEY_ENTER, 10, 13, 32):  # Enter or Space
            if current_idx == len(unique_categories):  # "New Category" selected
                new_category = get_input(stdscr, "Enter new category: ")
                if new_category in unique_categories:
                    category = new_category  # Assign to the existing category
                else:
                    unique_categories.append(new_category)
                    category = new_category
            else:
                category = unique_categories[page_idx * 10 + current_idx]  # Correct index for selected category
            
    # If Recurring Expense, Select Frequency
    frequency = "N/A"
    if exp_type == "Recurring":
        current_idx = 0
        while frequency == "N/A":
            draw_frequency_menu(stdscr, current_idx)
            key = stdscr.getch()
            if key == curses.KEY_UP and current_idx > 0:
                current_idx -= 1
            elif key == curses.KEY_DOWN and current_idx < 2:
                current_idx += 1
            elif key in (curses.KEY_ENTER, 10, 13, 32):  # Enter or Space
                frequency = ["Weekly", "Monthly", "Yearly"][current_idx]

    # Prompt for Description and Amount
    description = get_input(stdscr, "Enter description: ")
    amount = float(get_input(stdscr, "Enter amount: "))

    # Add Expense to File
    add_expense(file_name, exp_type, category, description, amount, frequency)


# Calculate Checking Balance
def calculate_checking_balance(file_name, stdscr):
    def draw_balance_menu(stdscr, current_idx):
        stdscr.clear()
        stdscr.addstr("Choose the frequency to view the checking balance:\n", curses.A_BOLD | curses.A_UNDERLINE)
        for idx, option in enumerate(["Weekly", "Monthly", "Yearly"]):
            if idx == current_idx:
                stdscr.addstr(f"> {option}\n", curses.A_REVERSE)
            else:
                stdscr.addstr(f"  {option}\n")
        stdscr.refresh()

    wb = openpyxl.load_workbook(file_name)
    settings = wb["Settings"]
    expenses = wb["Expenses"]

    salary = settings["B1"].value
    savings_ratio = settings["B2"].value
    checking_ratio = settings["B3"].value

    yearly_savings = salary * savings_ratio
    yearly_checking = salary * checking_ratio

    total_expenses = 0
    for row in expenses.iter_rows(min_row=2, values_only=True):
        amount = row[3]
        frequency = row[4]
        if frequency == "Weekly":
            total_expenses += amount * 52
        elif frequency == "Monthly":
            total_expenses += amount * 12
        elif frequency == "Yearly":
            total_expenses += amount

    checking_after_expenses = yearly_checking - total_expenses

    def balance_menu(stdscr):
        curses.curs_set(0)  # Hide cursor
        current_idx = 0
        frequencies = ["Weekly", "Monthly", "Yearly"]
        balance_values = [
            checking_after_expenses / 52,  # Weekly
            checking_after_expenses / 12,  # Monthly
            checking_after_expenses        # Yearly
        ]

        while True:
            draw_balance_menu(stdscr, current_idx)

            key = stdscr.getch()
            if key == curses.KEY_UP and current_idx > 0:
                current_idx -= 1
            elif key == curses.KEY_DOWN and current_idx < len(frequencies) - 1:
                current_idx += 1
            elif key in (curses.KEY_ENTER, 10, 13):  # Enter key
                period = frequencies[current_idx]
                balance = balance_values[current_idx]
                stdscr.clear()
                stdscr.addstr(f"Checking Balance ({period}):\n", curses.A_BOLD | curses.A_UNDERLINE)
                stdscr.addstr(f"Yearly Salary: ${salary}\n")
                stdscr.addstr(f"Savings: ${yearly_savings}\n")
                stdscr.addstr(f"Total Expenses: ${total_expenses}\n")
                stdscr.addstr(f"Checking Balance After Expenses ({period}): ${balance:.2f}\n")
                stdscr.addstr("\nPress any key to return to the menu.")
                stdscr.refresh()
                stdscr.getch()
                break

    balance_menu(stdscr)


# View Expenses with Row Numbers
def view_expenses(file_name, stdscr):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb["Expenses"]

    stdscr.clear()
    stdscr.addstr("Expenses:\n", curses.A_BOLD | curses.A_UNDERLINE)
    stdscr.addstr(f"{'Row':<5}{'Type':<15}{'Category':<15}{'Description':<25}{'Amount':<10}{'Frequency'}\n")
    stdscr.addstr("-" * 75 + "\n")
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        stdscr.addstr(f"{idx:<5}{row[0]:<15}{row[1]:<15}{row[2]:<25}{row[3]:<10.2f}{row[4]}\n")
    stdscr.addstr("\nPress any key to return to the menu.")
    stdscr.refresh()
    stdscr.getch()  # Wait for user input before returning

# Main Menu with Navigation
def budget_menu(file_name):
    menu_options = [
        "Set Yearly Salary",
        "Set Savings/Checking Ratios",
        "Add Expense",
        "Remove Expense",
        "View Expenses",
        "Calculate Checking Balance",
        "Exit",
    ]

    def draw_menu(stdscr, current_idx):
        stdscr.clear()
        stdscr.addstr("Budget Menu:\n", curses.A_BOLD | curses.A_UNDERLINE)
        for idx, option in enumerate(menu_options):
            if idx == current_idx:
                stdscr.addstr(f"> {option}\n", curses.A_REVERSE)
            else:
                stdscr.addstr(f"  {option}\n")
        stdscr.refresh()

    def handle_selection(stdscr, selected_idx):
        if selected_idx == 0:  # Set Yearly Salary
            salary = get_input(stdscr, "Enter your yearly salary: ")
            set_yearly_salary(file_name, float(salary))
        elif selected_idx == 1:  # Set Ratios
            savings_ratio = get_input(stdscr, "Enter savings ratio (0.0 - 1.0): ")
            if 0 <= float(savings_ratio) <= 1:
                set_ratios(file_name, float(savings_ratio))
            else:
                stdscr.addstr("Invalid ratio. Must be between 0.0 and 1.0.\n")
                stdscr.refresh()
        elif selected_idx == 2:  # Add Expense
            add_expense_menu(file_name, stdscr)
        elif selected_idx == 3:  # Remove Expense
            remove_expense(file_name, stdscr)
        elif selected_idx == 4:  # View Expenses
            view_expenses(file_name, stdscr)
        elif selected_idx == 5:  # Calculate Checking Balance
            calculate_checking_balance(file_name, stdscr)
        elif selected_idx == 6:  # Exit
            return False
        return True

    def curses_menu(stdscr):
        curses.curs_set(0)  # Hide cursor
        current_idx = 0
        while True:
            draw_menu(stdscr, current_idx)

            key = stdscr.getch()

            if key == curses.KEY_UP and current_idx > 0:
                current_idx -= 1
            elif key == curses.KEY_DOWN and current_idx < len(menu_options) - 1:
                current_idx += 1
            elif key in (curses.KEY_ENTER, 10, 13, 32):  # Enter or Space
                if not handle_selection(stdscr, current_idx):
                    break

    curses.wrapper(curses_menu)

# Main Program
if __name__ == "__main__":
    file_name = "Household_Budget.xlsx"
    initialize_budget_file(file_name)
    budget_menu(file_name)
